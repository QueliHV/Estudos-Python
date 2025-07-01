
from openpyxl import Workbook

#1. Criar um arquivo Excel
# Crie um arquivo chamado alunos.xlsx com os seguintes dados:
'''
Nom    Idade   Nota
Queli   39     9.5
Rafael  40     8.0
Alice    3    10.0
'''

wb = Workbook()   # Cria nova planilha
ws = wb.active  # Acessa a aba ativa 

# Cria colunas
ws.title = "Alunos"
ws['A1'] = "Nome"
ws['B1'] = "Idade"
ws['C1'] = "Nota"

ws.append = (["Queli", 39, 9.5])
ws.append = (["Rafael", 40, 8])
ws.append = (["Alice", 3, 10.0])


wb.save("modulo-06-Manipulacao-de-arquivos/arquivos/arquivo-xlsx.xlsx")







